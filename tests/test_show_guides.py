from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.guide_services import (
    create_guide_row,
    delete_guide_row,
    export_trade_show_guide_workbook,
    import_trade_show_guide_workbook,
    rebuild_trade_show_guides,
    update_guide_row,
)
from app.models import Show, ShowGuideRow, ShowStatus
from app.show_guides import build_guide_sheet_views


def make_show(**overrides) -> Show:
    payload = {
        "name": "ICA Car Wash Show",
        "event_date": date(2026, 5, 20),
        "place": "Las Vegas, NV",
        "source_url": "https://example.com/exhibitors",
        "run_offset_days": 14,
        "run_at": datetime(2026, 5, 6, 9, 0),
        "status": ShowStatus.ready_for_review.value,
        "latest_export_path": "",
        "company_count": 0,
        "failure_count": 0,
        "last_error": "",
    }
    payload.update(overrides)
    return Show(**payload)


def build_workbook_bytes() -> bytes:
    workbook = Workbook()
    company_sheet = workbook.active
    company_sheet.title = "Company Summary"
    company_sheet.append(
        [
            "Company Name",
            "Booth Number",
            "Booth Category",
            "Sales Team Size",
            "Customer Service Team Size",
            "Total Team Size",
            "Catalog Complexity (1-5)",
            "Sales Leader Name",
            "Sales Leader Role",
            "Sales Leader Email",
            "Sales Leader LinkedIn",
            "Source URL",
        ]
    )
    company_sheet.append(
        [
            "Fiserv",
            "3254",
            "3200s",
            932,
            2009,
            2941,
            2,
            "Robert Clarkson",
            "Chief Revenue Officer",
            "robert.clarkson@fiserv.com",
            "https://linkedin.com/in/robert",
            "https://example.com/fiserv",
        ]
    )

    booth_sheet = workbook.create_sheet("Booth Category Groups")
    booth_sheet.append(
        [
            "Booth Category",
            "Category Total Team Size",
            "Company Name",
            "Booth Number",
            "Sales Team Size",
            "Customer Service Team Size",
            "Total Team Size",
            "Catalog Complexity (1-5)",
            "Sales Leader Name",
            "Sales Leader Role",
            "Sales Leader Email",
            "Sales Leader LinkedIn",
            "Source URL",
        ]
    )
    booth_sheet.append(
        [
            "100s",
            20,
            "Special-Lite",
            "135",
            7,
            5,
            12,
            4,
            "Gary Wolf",
            "Business Development Manager",
            "gwolf@example.com",
            "https://linkedin.com/in/gary",
            "https://example.com/special-lite",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ShowGuideTests(unittest.TestCase):
    def setUp(self) -> None:
        self.engine = create_engine("sqlite:///:memory:", future=True)
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)

    def tearDown(self) -> None:
        self.engine.dispose()

    def test_create_update_delete_guide_row(self) -> None:
        with self.Session() as session:
            show = make_show()
            session.add(show)
            session.commit()

            row = create_guide_row(
                session,
                show=show,
                sheet_key="company_summary",
                payload={"company_name": "Fiserv", "booth_number": "3254"},
            )
            self.assertEqual(row.position, 0)

            update_guide_row(
                session,
                row=row,
                payload={"company_name": "Fiserv Updated", "booth_number": "3254"},
            )
            refreshed = session.scalar(select(ShowGuideRow).where(ShowGuideRow.id == row.id))
            assert refreshed is not None
            self.assertIn("Fiserv Updated", refreshed.values_json)

            delete_guide_row(session, row=row)
            deleted = session.scalar(select(ShowGuideRow).where(ShowGuideRow.id == row.id))
            self.assertIsNone(deleted)

    def test_build_guide_sheet_views_groups_rows_by_sheet(self) -> None:
        with self.Session() as session:
            show = make_show()
            session.add(show)
            session.commit()
            create_guide_row(session, show=show, sheet_key="company_summary", payload={"company_name": "Fiserv"})
            create_guide_row(session, show=show, sheet_key="booth_category_groups", payload={"booth_category": "100s"})
            session.refresh(show)

            views = build_guide_sheet_views(show)

            self.assertEqual(len(views), 2)
            self.assertEqual(views[0].definition.key, "company_summary")
            self.assertEqual(views[0].rows[0].values["company_name"], "Fiserv")
            self.assertEqual(views[1].rows[0].values["booth_category"], "100s")

    def test_rebuild_trade_show_guides_populates_both_sheets_from_export(self) -> None:
        with self.Session() as session:
            show = make_show(
                latest_export_path="/tmp/icff.csv",
            )
            session.add(show)
            session.commit()

            create_guide_row(session, show=show, sheet_key="company_summary", payload={"company_name": "Old Row"})

            with patch(
                "app.guide_services._load_company_rows",
                return_value=[
                    {
                        "company_name": "Acme",
                        "booth_number": "3254",
                        "website_url": "https://acme.com",
                    },
                    {
                        "company_name": "Beta",
                        "booth_number": "118",
                        "website_url": "",
                    },
                ],
            ):
                company_count, booth_count = rebuild_trade_show_guides(session, show=show)

            self.assertEqual(company_count, 2)
            self.assertEqual(booth_count, 2)
            rows = session.scalars(
                select(ShowGuideRow).where(ShowGuideRow.show_id == show.id).order_by(
                    ShowGuideRow.sheet_key.asc(),
                    ShowGuideRow.position.asc(),
                )
            ).all()
            self.assertEqual(len(rows), 4)
            rows_by_sheet = {
                sheet_key: [row for row in rows if row.sheet_key == sheet_key]
                for sheet_key in {"company_summary", "booth_category_groups"}
            }
            self.assertIn('"company_name": "Acme"', rows_by_sheet["company_summary"][0].values_json)
            self.assertIn('"booth_category": "3200s"', rows_by_sheet["company_summary"][0].values_json)
            self.assertIn(
                '"category_total_team_size": "1"',
                rows_by_sheet["booth_category_groups"][0].values_json,
            )

    def test_import_trade_show_guide_workbook_replaces_rows_from_excel(self) -> None:
        with self.Session() as session:
            show = make_show()
            session.add(show)
            session.commit()

            create_guide_row(session, show=show, sheet_key="company_summary", payload={"company_name": "Old Row"})

            counts = import_trade_show_guide_workbook(session, show=show, workbook_bytes=build_workbook_bytes())

            self.assertEqual(counts["company_summary"], 1)
            self.assertEqual(counts["booth_category_groups"], 1)
            rows = session.scalars(select(ShowGuideRow).where(ShowGuideRow.show_id == show.id)).all()
            self.assertEqual(len(rows), 2)
            self.assertTrue(any("Fiserv" in row.values_json for row in rows))

    def test_export_trade_show_guide_workbook_returns_xlsx_bytes(self) -> None:
        with self.Session() as session:
            show = make_show()
            session.add(show)
            session.commit()

            import_trade_show_guide_workbook(session, show=show, workbook_bytes=build_workbook_bytes())
            workbook_bytes = export_trade_show_guide_workbook(show)

            self.assertTrue(workbook_bytes.startswith(b"PK"))
            workbook = load_workbook(BytesIO(workbook_bytes))
            self.assertEqual(workbook.sheetnames, ["Company Summary", "Booth Category Groups"])


if __name__ == "__main__":
    unittest.main()
