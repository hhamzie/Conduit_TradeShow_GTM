from __future__ import annotations

import csv
from dataclasses import dataclass, field, replace
from datetime import datetime, timezone
import hashlib
import html
import io
import json
from collections.abc import Mapping
from pathlib import Path
import re
import shutil
import time
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

from app.config import get_settings


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
PIPEDRIVE_VISIBLE_TO_ENTIRE_COMPANY = 3
EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
PHONE_RE = re.compile(r"\+?\d[\d\s().-]{6,}\d")

TRADESHOW_FIELD_KEY = "dfcc62a9104fce98ed1ea6566a9bd82d002a6986"
INDUSTRY_FIELD_KEY = "6d24142b70f5c8b1b140c9b278bb7c8141fcb8bc"
PERSON_LINKEDIN_PROFILE_FIELD_KEY = "57139293fba428fc3ffebc689b0fecc628569aa0"
PERSON_LINKEDIN_URL_FIELD_KEY = "2e0318b3854d2c3996417e4b74e8ba0ea1f80813"
CLAY_PULL_PAGE_SIZE = 250
DEFAULT_PIPEDRIVE_PIPELINE_ID = 39
DEFAULT_PIPEDRIVE_STAGE_ID = 377
DEFAULT_DYNAMIC_EVENT_INDUSTRY_OPTION_ID = 277
UNRESOLVED_TRADESHOW_OPTION_ID = 0


@dataclass(frozen=True)
class RepMapping:
    slug: str
    name: str
    email: str
    pipedrive_user_id: int
    filename_terms: tuple[str, ...]


@dataclass(frozen=True)
class EventMapping:
    slug: str
    name: str
    filename_terms: tuple[str, ...]
    pipedrive_pipeline_id: int
    pipedrive_stage_id: int
    pipedrive_channel_id: str
    pipedrive_tradeshow_option_id: int
    pipedrive_industry_option_id: int


@dataclass(frozen=True)
class LeadSheetRow:
    company_name: str
    clay_row_id: str = ""
    person_name: str = ""
    first_name: str = ""
    last_name: str = ""
    email: str = ""
    phone: str = ""
    website_url: str = ""
    domain: str = ""
    linkedin_url: str = ""
    job_title: str = ""
    booth_number: str = ""
    location: str = ""
    conference: str = ""
    source_notes: str = ""
    sales_leader_name: str = ""
    sales_leader_email: str = ""
    sales_leader_phone: str = ""
    sales_leader_linkedin_url: str = ""
    sales_leader_title: str = ""
    rep_slug: str = ""
    rep_name: str = ""
    rep_email: str = ""
    rep_pipedrive_user_id: int | None = None
    source_file: str = ""
    source_row_number: int = 0
    enriched_status: str = ""
    extra: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class PipedriveContact:
    label: str
    name: str
    email: str = ""
    phone: str = ""
    linkedin_url: str = ""
    job_title: str = ""


@dataclass(frozen=True)
class SheetParseResult:
    path: Path
    rows: list[LeadSheetRow]
    rep: RepMapping
    event: EventMapping | None


@dataclass
class OperationSummary:
    scanned_files: int = 0
    parsed_rows: int = 0
    sent_to_clay: int = 0
    imported_to_pipedrive: int = 0
    updated_people: int = 0
    created_people: int = 0
    updated_orgs: int = 0
    created_orgs: int = 0
    updated_deals: int = 0
    created_deals: int = 0
    created_notes: int = 0
    skipped_rows: int = 0
    errors: list[str] = field(default_factory=list)


REPS: tuple[RepMapping, ...] = (
    RepMapping(
        slug="anand",
        name="Anand Prabhu",
        email="anand.prabhu@conduitcommerce.com",
        pipedrive_user_id=22329483,
        filename_terms=("anand", "prabhu", "anand_prabhu", "anand-prabhu"),
    ),
    RepMapping(
        slug="lea",
        name="Lea Skoumbakis",
        email="evangelia.skoumbakis@conduitcommerce.com",
        pipedrive_user_id=25200571,
        filename_terms=("lea", "evangelia", "skoumbakis"),
    ),
    RepMapping(
        slug="austin",
        name="Austin Weitman",
        email="austin.weitman@conduitcommerce.com",
        pipedrive_user_id=25188570,
        filename_terms=("austin", "weitman"),
    ),
    RepMapping(
        slug="gavin",
        name="Gavin Nagy",
        email="gavin.nagy@conduitcommercetech.com",
        pipedrive_user_id=25897289,
        filename_terms=("gavin", "nagy", "gavin_nagy", "gavin-nagy"),
    ),
    RepMapping(
        slug="hudson",
        name="Hudson Stedman",
        email="hudson@conduitcommercetech.com",
        pipedrive_user_id=23584737,
        filename_terms=("hudson", "stedman", "hudson_stedman", "hudson-stedman"),
    ),
    RepMapping(
        slug="hunter",
        name="Hunter Lee",
        email="hunter.lee@conduitcommerce.com",
        pipedrive_user_id=24521508,
        filename_terms=("hunter", "lee", "hunter_lee", "hunter-lee"),
    ),
    RepMapping(
        slug="john",
        name="John Yoon",
        email="john.yoon@conduitcommerce.com",
        pipedrive_user_id=25232735,
        filename_terms=("john", "yoon", "john_yoon", "john-yoon"),
    ),
    RepMapping(
        slug="noah",
        name="Noah Breen",
        email="noah.breen@conduit-commerce.com",
        pipedrive_user_id=24079506,
        filename_terms=("noah", "breen", "noah_breen", "noah-breen"),
    ),
)

EVENTS: tuple[EventMapping, ...] = (
    EventMapping(
        slug="infocomm",
        name="InfoComm Las Vegas",
        filename_terms=("infocomm", "info comm"),
        pipedrive_pipeline_id=DEFAULT_PIPEDRIVE_PIPELINE_ID,
        pipedrive_stage_id=DEFAULT_PIPEDRIVE_STAGE_ID,
        pipedrive_channel_id="InfoComm Las Vegas 2026",
        pipedrive_tradeshow_option_id=310,
        pipedrive_industry_option_id=311,
    ),
    EventMapping(
        slug="dallasmarket",
        name="Dallas Market",
        filename_terms=("dallasmarket", "dallas market", "dallas_market", "dallas-market"),
        pipedrive_pipeline_id=DEFAULT_PIPEDRIVE_PIPELINE_ID,
        pipedrive_stage_id=DEFAULT_PIPEDRIVE_STAGE_ID,
        pipedrive_channel_id="Dallas Market 2026",
        pipedrive_tradeshow_option_id=319,
        pipedrive_industry_option_id=DEFAULT_DYNAMIC_EVENT_INDUSTRY_OPTION_ID,
    ),
)

HEADER_ALIASES = {
    "company_name": (
        "company_name",
        "company name",
        "company",
        "organization",
        "organisation",
        "org",
        "account",
        "business",
        "exhibitor",
        "exhibitor name",
    ),
    "person_name": (
        "person_name",
        "person name",
        "contact_name",
        "contact name",
        "contact",
        "contact person",
        "poc",
        "full_name",
        "full name",
        "name",
    ),
    "first_name": ("first_name", "first name", "firstname", "first"),
    "last_name": ("last_name", "last name", "lastname", "last"),
    "email": ("email", "email_address", "email address", "work_email", "work email", "professional_email"),
    "phone": ("phone", "phone_number", "phone number", "mobile", "mobile_phone", "work_phone", "telephone"),
    "website_url": ("website_url", "website url", "website", "company_url", "company website", "url", "domain"),
    "domain": ("domain", "company domain", "company_domain", "official domain", "official domain domain"),
    "linkedin_url": (
        "linkedin_url",
        "linkedin url",
        "linkedin",
        "linkedin_profile",
        "linkedin profile",
        "linkedin profile url",
        "linked in profile url linkedin profile url",
    ),
    "job_title": ("job_title", "job title", "title", "role"),
    "booth_number": ("booth_number", "booth number", "booth", "stand", "stand number"),
    "location": ("location", "place", "city", "state", "hq location"),
    "conference": (
        "conference",
        "show",
        "show name",
        "event",
        "event name",
        "trade show",
        "tradeshow",
        "tradeshow name",
        "tradeshow name for emails",
        "source channel id",
    ),
    "rep": ("rep", "owner", "sales rep", "sales_rep", "assigned rep", "assigned_rep"),
    "rep_slug": ("rep_slug", "rep slug"),
    "rep_name": ("rep_name", "rep name", "owner name", "sales rep name"),
    "rep_email": ("rep_email", "rep email", "owner email", "sales rep email"),
    "rep_pipedrive_user_id": ("rep_pipedrive_user_id", "rep pipedrive user id", "owner pipedrive user id"),
    "source_file": ("source_file", "source file"),
    "source_notes": ("source_notes", "source notes", "notes", "note"),
    "clay_row_id": ("clay_row_id", "clay row id", "record id", "record_id", "row id"),
    "sales_leader_name": (
        "sales_leader_name",
        "sales leader name",
        "sales leader namae",
        "sales name",
        "sales person",
        "salesperson",
        "sales contact",
    ),
    "sales_leader_email": (
        "sales_leader_email",
        "sales leader email",
        "email data",
        "email - data",
        "find work email",
    ),
    "sales_leader_phone": (
        "sales_leader_phone",
        "sales leader phone",
        "sales leader mobile",
        "sales number",
        "mobile number",
        "find mobile number",
    ),
    "sales_leader_linkedin_url": (
        "sales_leader_linkedin_url",
        "sales leader linkedin",
        "sales leader linkedin url",
        "sales leader linkedin profile",
        "sales linkedin",
        "sales cooked linkedin",
    ),
    "sales_leader_title": ("sales_leader_title", "sales leader title", "sales leader role"),
    "enriched_status": ("enriched_status", "enrichment status", "status", "clay status"),
}


def setup_ingestion_dirs(root: Path | None = None) -> Path:
    root = root or get_settings().trade_show_ingestion_dir
    for dirname in ("ready", "processing", "processed", "failed", "enriched"):
        (root / dirname).mkdir(parents=True, exist_ok=True)
    return root


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _pick(row: dict[str, str], field: str) -> str:
    for alias in HEADER_ALIASES[field]:
        value = row.get(_normalize_header(alias), "")
        if value:
            return value
    return ""


def _pick_matching(row: dict[str, str], field: str, predicate) -> str:
    for alias in HEADER_ALIASES[field]:
        value = row.get(_normalize_header(alias), "")
        if value and predicate(value):
            return value
    return ""


def _pick_email(row: dict[str, str], field: str) -> str:
    for alias in HEADER_ALIASES[field]:
        value = row.get(_normalize_header(alias), "")
        if not value:
            continue
        match = EMAIL_RE.search(value)
        if match:
            return match.group(0)
    return ""


def _pick_phone(row: dict[str, str], field: str) -> str:
    for alias in HEADER_ALIASES[field]:
        value = row.get(_normalize_header(alias), "")
        if not value:
            continue
        match = PHONE_RE.search(value)
        if match:
            digits = re.sub(r"\D+", "", match.group(0))
            if len(digits) >= 7:
                return match.group(0).strip()
    return ""


def _pick_linkedin_url(row: dict[str, str], field: str) -> str:
    return _pick_matching(
        row,
        field,
        lambda value: "linkedin.com/" in value.lower() and value.lower().startswith(("http://", "https://")),
    )


def _split_name(name: str) -> tuple[str, str]:
    parts = name.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _domain_from_website(value: str) -> str:
    raw = value.strip()
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    host = (parsed.netloc or parsed.path.split("/", 1)[0]).lower()
    return host.removeprefix("www.")


def _int_value(value: object) -> int | None:
    try:
        if value in (None, ""):
            return None
        if isinstance(value, str) and value.strip().endswith(".0"):
            value = value.strip()[:-2]
        return int(value)
    except (TypeError, ValueError):
        return None


def _tokens_for_filename(path: Path) -> set[str]:
    return {token for token in re.split(r"[^a-z0-9]+", path.stem.lower()) if token}


GENERIC_EVENT_FILENAME_TOKENS = {
    "contacts",
    "contact",
    "enriched",
    "export",
    "lead",
    "leads",
    "list",
    "sheet",
    "upload",
}

EVENT_COMPOUND_SUFFIXES = ("market", "show", "expo")

EVENT_NAME_ACRONYMS = {
    "asd": "ASD",
    "f": "F",
    "b": "B",
    "kbis": "KBIS",
    "nra": "NRA",
    "ny": "NY",
    "nyc": "NYC",
}


def _split_compound_event_token(token: str) -> list[str]:
    for suffix in EVENT_COMPOUND_SUFFIXES:
        if token.endswith(suffix) and len(token) > len(suffix) + 2:
            return [token[: -len(suffix)], suffix]
    return [token]


def _split_filename_event_token(token: str) -> list[str]:
    camel_parts = re.findall(r"[A-Z]+(?=[A-Z][a-z]|$)|[A-Z]?[a-z]+|[0-9]+", token)
    if len(camel_parts) > 1:
        parts: list[str] = []
        for part in camel_parts:
            parts.extend(_split_compound_event_token(part.lower()))
        return parts
    return _split_compound_event_token(token.lower())


def _title_event_token(token: str) -> str:
    return EVENT_NAME_ACRONYMS.get(token.lower(), token.capitalize())


def _event_name_from_tokens(tokens: list[str]) -> str:
    return " ".join(_title_event_token(token) for token in tokens if token)


def _clean_event_name(value: str) -> str:
    cleaned = _clean(value)
    cleaned = re.sub(r"\.[A-Za-z0-9]+$", "", cleaned)
    cleaned = re.sub(r"\b20\d{2}\b", "", cleaned)
    cleaned = re.sub(r"[_-]+", " ", cleaned)
    cleaned = _clean(cleaned)
    if not cleaned:
        return ""
    tokens = [
        split
        for token in re.split(r"[^A-Za-z0-9]+", cleaned)
        if token
        for split in _split_compound_event_token(token.lower())
    ]
    return _event_name_from_tokens(tokens)


def _event_name_from_filename(path: Path, rep: RepMapping | None = None) -> str:
    rep_tokens = set()
    if rep is not None:
        rep_tokens.update(_normalize_key(term) for term in rep.filename_terms)
        rep_tokens.add(rep.slug)
    tokens: list[str] = []
    for raw_token in re.split(r"[^A-Za-z0-9]+", path.stem):
        if not raw_token:
            continue
        normalized_token = raw_token.lower()
        if normalized_token in rep_tokens or normalized_token in GENERIC_EVENT_FILENAME_TOKENS:
            continue
        tokens.extend(_split_filename_event_token(raw_token))
    return _event_name_from_tokens(tokens)


def _dynamic_event_mapping(event_name: str) -> EventMapping | None:
    cleaned_name = _clean_event_name(event_name)
    if not cleaned_name:
        return None
    slug = _normalize_key(cleaned_name)
    return EventMapping(
        slug=slug,
        name=cleaned_name,
        filename_terms=(slug, _normalize_header(cleaned_name)),
        pipedrive_pipeline_id=DEFAULT_PIPEDRIVE_PIPELINE_ID,
        pipedrive_stage_id=DEFAULT_PIPEDRIVE_STAGE_ID,
        pipedrive_channel_id=f"{cleaned_name} {datetime.now(timezone.utc).year}",
        pipedrive_tradeshow_option_id=UNRESOLVED_TRADESHOW_OPTION_ID,
        pipedrive_industry_option_id=DEFAULT_DYNAMIC_EVENT_INDUSTRY_OPTION_ID,
    )


def _event_option_match_key(event_name: str) -> str:
    return _normalize_header(_clean_event_name(event_name))


def infer_rep(path: Path) -> RepMapping:
    tokens = _tokens_for_filename(path)
    raw_name = path.stem.lower()
    for rep in REPS:
        if any(term in tokens or term in raw_name for term in rep.filename_terms):
            return rep
    raise ValueError(f"Could not infer rep from filename: {path.name}")


def infer_rep_from_rows(path: Path, rows: list[dict[str, str]]) -> RepMapping:
    try:
        return infer_rep(path)
    except ValueError:
        pass

    for row in rows[:25]:
        rep_id = _int_value(_pick(row, "rep_pipedrive_user_id"))
        rep_slug = _pick(row, "rep_slug").strip().lower()
        rep_name = _pick(row, "rep_name").strip().lower()
        rep_email = _pick(row, "rep_email").strip().lower()
        source_file = _pick(row, "source_file")
        for rep in REPS:
            if rep_id == rep.pipedrive_user_id:
                return rep
            if rep_slug == rep.slug:
                return rep
            if rep_email == rep.email.lower():
                return rep
            if rep_name and rep_name == rep.name.lower():
                return rep
            if source_file and any(term in source_file.lower() for term in rep.filename_terms):
                return rep
    raise ValueError(f"Could not infer rep from filename or rows: {path.name}")


def infer_event(path: Path, rows: list[dict[str, str]], *, rep: RepMapping | None = None) -> EventMapping | None:
    haystack_parts = [path.stem.lower()]
    for row in rows[:10]:
        haystack_parts.extend([
            _pick(row, "conference"),
            _pick(row, "source_file"),
        ])
    haystack = " ".join(haystack_parts).lower()
    for event in EVENTS:
        if any(term in haystack for term in event.filename_terms):
            return event

    for row in rows[:25]:
        event_name = _pick(row, "conference")
        dynamic_event = _dynamic_event_mapping(event_name)
        if dynamic_event is not None:
            return dynamic_event

    for row in rows[:25]:
        source_file = _pick(row, "source_file")
        if not source_file:
            continue
        dynamic_event = _dynamic_event_mapping(_event_name_from_filename(Path(source_file), rep))
        if dynamic_event is not None:
            return dynamic_event

    dynamic_event = _dynamic_event_mapping(_event_name_from_filename(path, rep))
    if dynamic_event is not None:
        return dynamic_event
    return None


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.DictReader(csv_file)
        headers = [_normalize_header(header) for header in (reader.fieldnames or [])]
        rows: list[dict[str, str]] = []
        for raw_row in reader:
            normalized: dict[str, str] = {}
            for index, original_key in enumerate(raw_row.keys()):
                header = headers[index] if index < len(headers) else _normalize_header(original_key)
                normalized[header] = _clean(raw_row.get(original_key))
            rows.append(normalized)
        return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers: list[str] | None = None
    records: list[dict[str, str]] = []
    for values in rows_iter:
        cells = [_clean(value) for value in values]
        if headers is None:
            if not any(cells):
                continue
            headers = [_normalize_header(value) for value in cells]
            continue
        record = {
            headers[index]: _clean(value)
            for index, value in enumerate(cells)
            if index < len(headers) and headers[index]
        }
        if any(record.values()):
            records.append(record)
    workbook.close()
    return records


def read_sheet_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.name}")


def normalize_rows(path: Path, raw_rows: list[dict[str, str]], *, limit: int | None = None) -> SheetParseResult:
    rep = infer_rep_from_rows(path, raw_rows)
    event = infer_event(path, raw_rows, rep=rep)
    normalized_rows: list[LeadSheetRow] = []
    for row_number, row in enumerate(raw_rows, start=2):
        company_name = _pick(row, "company_name")
        if not company_name:
            continue
        person_name = _pick(row, "person_name")
        first_name = _pick(row, "first_name")
        last_name = _pick(row, "last_name")
        if person_name and not (first_name or last_name):
            first_name, last_name = _split_name(person_name)
        if not person_name and (first_name or last_name):
            person_name = " ".join(part for part in (first_name, last_name) if part)
        website_url = _pick(row, "website_url")
        domain = _pick(row, "domain") or _domain_from_website(website_url)
        rep_id = _int_value(_pick(row, "rep_pipedrive_user_id")) or rep.pipedrive_user_id
        known_headers = {_normalize_header(alias) for aliases in HEADER_ALIASES.values() for alias in aliases}
        extra = {
            _normalize_key(key): value
            for key, value in row.items()
            if value and key not in known_headers
        }
        normalized_rows.append(
            LeadSheetRow(
                company_name=company_name,
                clay_row_id=_pick(row, "clay_row_id"),
                person_name=person_name,
                first_name=first_name,
                last_name=last_name,
                email=_pick_email(row, "email").lower(),
                phone=_pick_phone(row, "phone"),
                website_url=website_url,
                domain=domain,
                linkedin_url=_pick_linkedin_url(row, "linkedin_url"),
                job_title=_pick(row, "job_title"),
                booth_number=_pick(row, "booth_number"),
                location=_pick(row, "location"),
                conference=_pick(row, "conference") or (event.name if event else ""),
                source_notes=_pick(row, "source_notes"),
                sales_leader_name=_pick(row, "sales_leader_name"),
                sales_leader_email=_pick_email(row, "sales_leader_email").lower(),
                sales_leader_phone=_pick_phone(row, "sales_leader_phone"),
                sales_leader_linkedin_url=_pick_linkedin_url(row, "sales_leader_linkedin_url"),
                sales_leader_title=_pick(row, "sales_leader_title"),
                rep_slug=_pick(row, "rep_slug") or rep.slug,
                rep_name=_pick(row, "rep_name") or rep.name,
                rep_email=_pick(row, "rep_email") or rep.email,
                rep_pipedrive_user_id=rep_id,
                source_file=_pick(row, "source_file") or path.name,
                source_row_number=row_number,
                enriched_status=_pick(row, "enriched_status").lower(),
                extra=extra,
            )
        )
        if limit is not None and len(normalized_rows) >= limit:
            break
    return SheetParseResult(path=path, rows=normalized_rows, rep=rep, event=event)


def normalize_sheet(path: Path, *, limit: int | None = None) -> SheetParseResult:
    return normalize_rows(path, read_sheet_rows(path), limit=limit)


def row_to_payload(row: LeadSheetRow) -> dict[str, object]:
    payload = {
        "company_name": row.company_name,
        "clay_row_id": row.clay_row_id,
        "person_name": row.person_name,
        "first_name": row.first_name,
        "last_name": row.last_name,
        "email": row.email,
        "phone": row.phone,
        "website": row.website_url,
        "domain": row.domain,
        "linkedin_url": row.linkedin_url,
        "job_title": row.job_title,
        "booth_number": row.booth_number,
        "location": row.location,
        "conference": row.conference,
        "source_notes": row.source_notes,
        "sales_leader_name": row.sales_leader_name,
        "sales_leader_email": row.sales_leader_email,
        "sales_leader_phone": row.sales_leader_phone,
        "sales_leader_linkedin_url": row.sales_leader_linkedin_url,
        "sales_leader_title": row.sales_leader_title,
        "rep_slug": row.rep_slug,
        "rep_name": row.rep_name,
        "rep_email": row.rep_email,
        "rep_pipedrive_user_id": row.rep_pipedrive_user_id,
        "source_file": row.source_file,
        "source_row_number": row.source_row_number,
        "sent_to_clay_at": datetime.now(timezone.utc).isoformat(),
    }
    payload.update({f"source_{key}": value for key, value in row.extra.items() if value})
    return {key: value for key, value in payload.items() if value not in ("", None)}


def primary_contact_for_row(row: LeadSheetRow) -> PipedriveContact:
    return PipedriveContact(
        label="Initial trade-show contact",
        name=row.person_name,
        email=row.email,
        phone=row.phone,
        linkedin_url=row.linkedin_url,
        job_title=row.job_title,
    )


NEGATIVE_ENRICHMENT_PATTERNS = (
    "could not determine",
    "descriptive statement",
    "identified top sales leader in the corpus",
    "missing input",
    "matches provided person name",
    "no alternate person",
    "no appropriate",
    "no valid",
    "no output",
    "not found",
    "no qualifying",
    "no phone found",
    "invalid input",
    "does not contain a valid",
    "not a valid",
    "available sources do not contain",
)

NEGATIVE_ENRICHMENT_VALUES = {"none", "null", "n/a", "na", "-"}


def _clean_enriched_contact_value(value: str) -> str:
    cleaned = _clean(value)
    lowered = cleaned.lower()
    if lowered in NEGATIVE_ENRICHMENT_VALUES:
        return ""
    if any(pattern in lowered for pattern in NEGATIVE_ENRICHMENT_PATTERNS):
        return ""
    return cleaned


def sales_leader_contact_for_row(row: LeadSheetRow) -> PipedriveContact:
    return PipedriveContact(
        label="Sales leader",
        name=_clean_enriched_contact_value(row.sales_leader_name),
        email=_clean_enriched_contact_value(row.sales_leader_email),
        phone=_clean_enriched_contact_value(row.sales_leader_phone),
        linkedin_url=_clean_enriched_contact_value(row.sales_leader_linkedin_url),
        job_title=_clean_enriched_contact_value(row.sales_leader_title),
    )


def contact_has_data(contact: PipedriveContact) -> bool:
    return any([contact.name, contact.email, contact.phone, contact.linkedin_url])


def contacts_look_same(left: PipedriveContact, right: PipedriveContact) -> bool:
    if left.email and right.email and left.email.lower() == right.email.lower():
        return True
    if left.linkedin_url and right.linkedin_url and left.linkedin_url.lower() == right.linkedin_url.lower():
        return True
    return bool(left.name and right.name and left.name.strip().lower() == right.name.strip().lower())


def _contact_note_line(contact: PipedriveContact) -> str:
    parts = [contact.label]
    details = []
    if contact.name:
        details.append(contact.name)
    if contact.job_title:
        details.append(contact.job_title)
    if contact.email:
        details.append(contact.email)
    if contact.phone:
        details.append(contact.phone)
    if contact.linkedin_url:
        details.append(contact.linkedin_url)
    return ": ".join([parts[0], " | ".join(details)]) if details else ""


def build_import_note(row: LeadSheetRow) -> str:
    lines = ["Trade show lead import"]
    source_bits = []
    if row.source_file:
        source_bits.append(row.source_file)
    if row.source_row_number:
        source_bits.append(f"row {row.source_row_number}")
    if row.clay_row_id:
        source_bits.append(f"Clay row {row.clay_row_id}")
    if source_bits:
        lines.append(f"Source: {' | '.join(source_bits)}")
    if row.conference:
        lines.append(f"Conference: {row.conference}")
    primary_line = _contact_note_line(primary_contact_for_row(row))
    if primary_line:
        lines.append(primary_line)
    sales_leader_line = _contact_note_line(sales_leader_contact_for_row(row))
    if sales_leader_line and not contacts_look_same(primary_contact_for_row(row), sales_leader_contact_for_row(row)):
        lines.append(sales_leader_line)
    if row.source_notes:
        lines.append(f"Source notes: {row.source_notes}")
    if len(lines) == 1:
        return ""
    return "<br>".join(html.escape(line) for line in lines)


def _request_json(
    method: str,
    url: str,
    *,
    headers: dict[str, str] | None = None,
    params: dict[str, object] | None = None,
    payload: object | None = None,
    timeout: float = 45.0,
) -> tuple[int, object]:
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        response = client.request(method, url, headers=headers, params=params, json=payload)
        try:
            body: object = response.json()
        except ValueError:
            body = response.text
        response.raise_for_status()
        return response.status_code, body


def send_rows_to_clay(rows: list[LeadSheetRow], *, dry_run: bool) -> OperationSummary:
    settings = get_settings()
    if not settings.clay_webhook_url:
        raise ValueError("CLAY_WEBHOOK_URL is not configured.")
    headers: dict[str, str] = {}
    if settings.clay_webhook_auth_header and settings.clay_webhook_auth_value:
        headers[settings.clay_webhook_auth_header] = settings.clay_webhook_auth_value

    sent = 0
    for row in rows:
        if dry_run:
            continue
        _request_json("POST", settings.clay_webhook_url, headers=headers, payload=row_to_payload(row), timeout=30.0)
        sent += 1
        time.sleep(0.08)
    return OperationSummary(parsed_rows=len(rows), sent_to_clay=sent)


class PipedriveClient:
    def __init__(self) -> None:
        settings = get_settings()
        self.base_url = settings.pipedrive_base_url
        self.api_token = settings.pipedrive_api_token.strip()
        self.headers = {"x-api-token": self.api_token}
        if not self.api_token:
            raise ValueError("PIPEDRIVE_API_TOKEN is not configured.")
        self.api_root_url = self.base_url.rsplit("/v1", 1)[0]

    def request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, object] | None = None,
        payload: dict[str, object] | None = None,
    ) -> dict[str, object]:
        _status, body = _request_json(
            method,
            f"{self.base_url}{path}",
            headers=self.headers,
            params=params,
            payload=payload,
        )
        if isinstance(body, dict):
            return body
        return {}

    def request_v2(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, object] | None = None,
        payload: object | None = None,
    ) -> dict[str, object]:
        query = dict(params or {})
        query["api_token"] = self.api_token
        _status, body = _request_json(
            method,
            f"{self.api_root_url}/api/v2{path}",
            headers={"Accept": "application/json"},
            params=query,
            payload=payload,
        )
        if isinstance(body, dict):
            return body
        return {}

    def deal_field_options(self, field_key: str) -> list[dict[str, object]]:
        body = self.request_v2("GET", f"/dealFields/{field_key}")
        data = body.get("data") if isinstance(body, dict) else None
        options = data.get("options") if isinstance(data, dict) else None
        return [option for option in options or [] if isinstance(option, dict)]

    def ensure_tradeshow_option(self, event_name: str, *, dry_run: bool) -> int:
        event_key = _event_option_match_key(event_name)
        for option in self.deal_field_options(TRADESHOW_FIELD_KEY):
            option_id = _int_value(option.get("id"))
            label = _clean(option.get("label"))
            if option_id is not None and _event_option_match_key(label) == event_key:
                return option_id
        if dry_run:
            return UNRESOLVED_TRADESHOW_OPTION_ID
        body = self.request_v2(
            "POST",
            f"/dealFields/{TRADESHOW_FIELD_KEY}/options",
            payload=[{"label": event_name}],
        )
        data = body.get("data") if isinstance(body, dict) else None
        first_option = data[0] if isinstance(data, list) and data else None
        option_id = _int_value(first_option.get("id") if isinstance(first_option, dict) else None)
        if option_id is None:
            raise ValueError(f"Pipedrive did not return a tradeshow option id for {event_name}.")
        return option_id

    def search_person_by_email(self, email: str) -> dict[str, object] | None:
        if not email:
            return None
        body = self.request(
            "GET",
            "/persons/search",
            params={"term": email, "fields": "email", "exact_match": "true", "limit": 10},
        )
        items = (body.get("data") or {}).get("items") if isinstance(body.get("data"), dict) else []
        for item in items or []:
            candidate = item.get("item") or {}
            emails = [str(value).lower() for value in candidate.get("emails") or []]
            if email.lower() in emails:
                return candidate
        return (items or [{}])[0].get("item") if items else None

    def search_person_by_name(self, name: str, org_id: int | None) -> dict[str, object] | None:
        if not name:
            return None
        body = self.request(
            "GET",
            "/persons/search",
            params={"term": name, "fields": "name", "exact_match": "true", "limit": 10},
        )
        items = (body.get("data") or {}).get("items") if isinstance(body.get("data"), dict) else []
        normalized = name.strip().lower()
        fallback: dict[str, object] | None = None
        for item in items or []:
            candidate = item.get("item") or {}
            if str(candidate.get("name") or "").strip().lower() != normalized:
                continue
            fallback = fallback or candidate
            organization = candidate.get("organization") or {}
            candidate_org_id = _int_value(organization.get("id") if isinstance(organization, dict) else None)
            if org_id is not None and candidate_org_id == org_id:
                return candidate
        return fallback

    def search_org_by_name(self, name: str) -> dict[str, object] | None:
        if not name:
            return None
        body = self.request(
            "GET",
            "/organizations/search",
            params={"term": name, "fields": "name", "exact_match": "true", "limit": 10},
        )
        items = (body.get("data") or {}).get("items") if isinstance(body.get("data"), dict) else []
        normalized = name.strip().lower()
        for item in items or []:
            candidate = item.get("item") or {}
            if str(candidate.get("name") or "").strip().lower() == normalized:
                return candidate
        return (items or [{}])[0].get("item") if items else None

    def search_open_deal(self, row: LeadSheetRow, event: EventMapping) -> dict[str, object] | None:
        body = self.request(
            "GET",
            "/deals/search",
            params={"term": row.company_name, "fields": "title", "exact_match": "true", "limit": 10},
        )
        items = (body.get("data") or {}).get("items") if isinstance(body.get("data"), dict) else []
        for item in items or []:
            candidate = item.get("item") or {}
            owner = candidate.get("owner") or {}
            pipeline = candidate.get("pipeline") or {}
            deal_id = _int_value(candidate.get("id"))
            if (
                str(candidate.get("status") or "") == "open"
                and str(candidate.get("title") or "").strip().lower() == row.company_name.strip().lower()
                and int(owner.get("id") or 0) == int(row.rep_pipedrive_user_id or 0)
                and int(pipeline.get("id") or 0) == event.pipedrive_pipeline_id
                and deal_id is not None
            ):
                detail = self.request("GET", f"/deals/{deal_id}").get("data") or {}
                if (
                    _int_value(detail.get(TRADESHOW_FIELD_KEY)) == event.pipedrive_tradeshow_option_id
                    or str(detail.get("channel_id") or "") == event.pipedrive_channel_id
                ):
                    return candidate
        return None

    def upsert_organization(self, row: LeadSheetRow, *, dry_run: bool) -> tuple[int | None, bool]:
        existing = self.search_org_by_name(row.company_name)
        payload = {
            "name": row.company_name,
            "owner_id": row.rep_pipedrive_user_id,
            "visible_to": PIPEDRIVE_VISIBLE_TO_ENTIRE_COMPANY,
        }
        if existing:
            org_id = int(existing["id"])
            if not dry_run:
                self.request("PUT", f"/organizations/{org_id}", payload=payload)
            return org_id, False
        if dry_run:
            return None, True
        body = self.request("POST", "/organizations", payload=payload)
        return int((body.get("data") or {}).get("id")), True

    def upsert_contact(
        self,
        contact: PipedriveContact,
        row: LeadSheetRow,
        org_id: int | None,
        *,
        dry_run: bool,
    ) -> tuple[int | None, bool]:
        if not contact_has_data(contact):
            return None, False
        existing = self.search_person_by_email(contact.email) if contact.email else None
        if existing is None and contact.name:
            existing = self.search_person_by_name(contact.name, org_id)
        payload: dict[str, object] = {
            "name": contact.name or contact.email or contact.phone or contact.linkedin_url or row.company_name,
            "owner_id": row.rep_pipedrive_user_id,
            "visible_to": PIPEDRIVE_VISIBLE_TO_ENTIRE_COMPANY,
        }
        if org_id is not None:
            payload["org_id"] = org_id
        if contact.email:
            payload["email"] = [{"value": contact.email, "primary": True, "label": "work"}]
        if contact.phone:
            payload["phone"] = [{"value": contact.phone, "primary": True, "label": "work"}]
        if contact.job_title:
            payload["job_title"] = contact.job_title
        if contact.linkedin_url:
            payload[PERSON_LINKEDIN_PROFILE_FIELD_KEY] = contact.linkedin_url
            payload[PERSON_LINKEDIN_URL_FIELD_KEY] = contact.linkedin_url
        if existing:
            person_id = int(existing["id"])
            if not dry_run:
                self.request("PUT", f"/persons/{person_id}", payload=payload)
            return person_id, False
        if dry_run:
            return None, True
        body = self.request("POST", "/persons", payload=payload)
        return int((body.get("data") or {}).get("id")), True

    def upsert_person(self, row: LeadSheetRow, org_id: int | None, *, dry_run: bool) -> tuple[int | None, bool]:
        return self.upsert_contact(primary_contact_for_row(row), row, org_id, dry_run=dry_run)

    def upsert_sales_leader(self, row: LeadSheetRow, org_id: int | None, *, dry_run: bool) -> tuple[int | None, bool]:
        primary = primary_contact_for_row(row)
        sales_leader = sales_leader_contact_for_row(row)
        if contacts_look_same(primary, sales_leader):
            return None, False
        return self.upsert_contact(sales_leader, row, org_id, dry_run=dry_run)

    def upsert_deal(
        self,
        row: LeadSheetRow,
        event: EventMapping,
        org_id: int | None,
        person_id: int | None,
        *,
        dry_run: bool,
    ) -> tuple[int | None, bool]:
        existing = self.search_open_deal(row, event)
        payload: dict[str, object] = {
            "title": row.company_name,
            "user_id": row.rep_pipedrive_user_id,
            "pipeline_id": event.pipedrive_pipeline_id,
            "stage_id": event.pipedrive_stage_id,
            "value": 0,
            "visible_to": PIPEDRIVE_VISIBLE_TO_ENTIRE_COMPANY,
            "channel_id": event.pipedrive_channel_id,
            TRADESHOW_FIELD_KEY: event.pipedrive_tradeshow_option_id,
            INDUSTRY_FIELD_KEY: event.pipedrive_industry_option_id,
        }
        if org_id is not None:
            payload["org_id"] = org_id
        if person_id is not None:
            payload["person_id"] = person_id
        if existing:
            deal_id = int(existing["id"])
            if not dry_run:
                self.request("PUT", f"/deals/{deal_id}", payload=payload)
            return deal_id, False
        if dry_run:
            return None, True
        payload["origin"] = "API"
        body = self.request("POST", "/deals", payload=payload)
        return int((body.get("data") or {}).get("id")), True

    def create_import_note(
        self,
        row: LeadSheetRow,
        *,
        deal_id: int | None,
        org_id: int | None,
        primary_person_id: int | None,
        sales_leader_person_id: int | None,
        dry_run: bool,
    ) -> bool:
        note_content = build_import_note(row)
        if not note_content:
            return False
        payload: dict[str, object] = {"content": note_content}
        if deal_id is not None:
            payload["deal_id"] = deal_id
        if org_id is not None:
            payload["org_id"] = org_id
        if sales_leader_person_id is not None:
            payload["person_id"] = sales_leader_person_id
        elif primary_person_id is not None:
            payload["person_id"] = primary_person_id
        if dry_run:
            return True
        self.request("POST", "/notes", payload=payload)
        return True


def import_rows_to_pipedrive(
    rows: list[LeadSheetRow],
    event: EventMapping | None,
    *,
    dry_run: bool,
    client: PipedriveClient | None = None,
    create_notes: bool = True,
) -> OperationSummary:
    if event is None:
        raise ValueError("Could not infer event mapping for Pipedrive import.")
    client = client or PipedriveClient()
    if event.pipedrive_tradeshow_option_id == UNRESOLVED_TRADESHOW_OPTION_ID:
        option_id = client.ensure_tradeshow_option(event.name, dry_run=dry_run)
        event = replace(event, pipedrive_tradeshow_option_id=option_id)
    summary = OperationSummary(parsed_rows=len(rows))
    for row in rows:
        if row.enriched_status and row.enriched_status not in {"ready", "complete", "completed"}:
            summary.skipped_rows += 1
            continue
        row_event = replace(event, pipedrive_channel_id=row.conference) if row.conference else event
        try:
            org_id, org_created = client.upsert_organization(row, dry_run=dry_run)
            person_id, person_created = client.upsert_person(row, org_id, dry_run=dry_run)
            sales_leader_id, sales_leader_created = client.upsert_sales_leader(row, org_id, dry_run=dry_run)
            deal_person_id = person_id or sales_leader_id
            deal_id, deal_created = client.upsert_deal(row, row_event, org_id, deal_person_id, dry_run=dry_run)
            note_created = False
            if create_notes:
                note_created = client.create_import_note(
                    row,
                    deal_id=deal_id,
                    org_id=org_id,
                    primary_person_id=person_id,
                    sales_leader_person_id=sales_leader_id,
                    dry_run=dry_run,
                )
            summary.imported_to_pipedrive += 1
            summary.created_orgs += int(org_created)
            summary.updated_orgs += int(not org_created)
            if person_id is not None or person_created:
                summary.created_people += int(person_created)
                summary.updated_people += int(not person_created)
            if sales_leader_id is not None or sales_leader_created:
                summary.created_people += int(sales_leader_created)
                summary.updated_people += int(not sales_leader_created)
            summary.created_deals += int(deal_created)
            summary.updated_deals += int(not deal_created)
            summary.created_notes += int(note_created)
        except (httpx.HTTPError, ValueError, TypeError) as exc:
            summary.errors.append(f"{row.company_name}: {exc}")
    return summary


def ready_files(root: Path | None = None) -> list[Path]:
    root = root or get_settings().trade_show_ingestion_dir
    ready_dir = root / "ready"
    if not ready_dir.exists():
        return []
    return sorted(
        path
        for path in ready_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


def enriched_files(root: Path | None = None) -> list[Path]:
    root = root or get_settings().trade_show_ingestion_dir
    enriched_dir = root / "enriched"
    if not enriched_dir.exists():
        return []
    return sorted(
        path
        for path in enriched_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


def _unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return path.with_name(f"{stem}_{timestamp}{suffix}")


def merge_summaries(*summaries: OperationSummary) -> OperationSummary:
    merged = OperationSummary()
    for summary in summaries:
        merged.scanned_files += summary.scanned_files
        merged.parsed_rows += summary.parsed_rows
        merged.sent_to_clay += summary.sent_to_clay
        merged.imported_to_pipedrive += summary.imported_to_pipedrive
        merged.updated_people += summary.updated_people
        merged.created_people += summary.created_people
        merged.updated_orgs += summary.updated_orgs
        merged.created_orgs += summary.created_orgs
        merged.updated_deals += summary.updated_deals
        merged.created_deals += summary.created_deals
        merged.created_notes += summary.created_notes
        merged.skipped_rows += summary.skipped_rows
        merged.errors.extend(summary.errors)
    return merged


def process_ready_files(
    *,
    root: Path | None = None,
    dry_run: bool = True,
    limit: int | None = None,
) -> OperationSummary:
    root = setup_ingestion_dirs(root)
    summary = OperationSummary()
    for source_path in ready_files(root):
        summary.scanned_files += 1
        processing_path = root / "processing" / source_path.name
        final_path = root / "processed" / source_path.name
        failed_path = root / "failed" / source_path.name
        active_path = source_path
        try:
            if not dry_run:
                shutil.move(str(source_path), str(processing_path))
                active_path = processing_path
            parsed = normalize_sheet(active_path, limit=limit)
            clay_result = send_rows_to_clay(parsed.rows, dry_run=dry_run)
            summary.parsed_rows += clay_result.parsed_rows
            summary.sent_to_clay += clay_result.sent_to_clay
            if not dry_run:
                shutil.move(str(active_path), str(final_path))
        except Exception as exc:
            summary.errors.append(f"{source_path.name}: {exc}")
            if not dry_run and active_path.exists():
                shutil.move(str(active_path), str(failed_path))
    return summary


def process_enriched_files(
    *,
    root: Path | None = None,
    dry_run: bool = True,
    limit: int | None = None,
    client: PipedriveClient | None = None,
) -> OperationSummary:
    root = setup_ingestion_dirs(root)
    summary = OperationSummary()
    for source_path in enriched_files(root):
        summary.scanned_files += 1
        archived_name = f"enriched__{source_path.name}"
        processing_path = _unique_destination(root / "processing" / archived_name)
        final_path = _unique_destination(root / "processed" / archived_name)
        failed_path = _unique_destination(root / "failed" / archived_name)
        active_path = source_path
        try:
            if not dry_run:
                shutil.move(str(source_path), str(processing_path))
                active_path = processing_path
            parsed = normalize_sheet(active_path, limit=limit)
            pipedrive_result = import_rows_to_pipedrive(
                parsed.rows,
                parsed.event,
                dry_run=dry_run,
                client=client,
            )
            summary = merge_summaries(summary, pipedrive_result)
            if not dry_run:
                shutil.move(str(active_path), str(final_path))
        except Exception as exc:
            summary.errors.append(f"{source_path.name}: {exc}")
            if not dry_run and active_path.exists():
                shutil.move(str(active_path), str(failed_path))
    return summary


def _clay_import_state_path(root: Path) -> Path:
    return root / "processed" / ".trade_show_pipedrive_imported_clay_rows.json"


def _clay_webhook_import_state_path(root: Path) -> Path:
    return root / "processed" / ".trade_show_pipedrive_imported_webhook_rows.json"


def _load_imported_clay_rows(root: Path) -> set[str]:
    return _load_imported_row_ids(_clay_import_state_path(root))


def _load_imported_clay_webhook_rows(root: Path) -> set[str]:
    return _load_imported_row_ids(_clay_webhook_import_state_path(root))


def _load_imported_row_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return set()
    if isinstance(payload, list):
        return {str(value) for value in payload if value}
    if isinstance(payload, dict):
        rows = payload.get("imported_clay_row_ids")
        if isinstance(rows, list):
            return {str(value) for value in rows if value}
    return set()


def _save_imported_clay_rows(root: Path, imported_row_ids: set[str]) -> None:
    _save_imported_row_ids(_clay_import_state_path(root), imported_row_ids)


def _save_imported_clay_webhook_rows(root: Path, imported_row_ids: set[str]) -> None:
    _save_imported_row_ids(_clay_webhook_import_state_path(root), imported_row_ids)


def _clay_local_export_path(root: Path) -> Path:
    return root / "enriched" / "clay_post_tradeshow_linkedin_enrichment.csv"


def _clay_table_snapshot_path(root: Path, table_id: str) -> Path:
    safe_table_id = _normalize_key(table_id) or "clay_table"
    return root / "processed" / f"clay_table_{safe_table_id}_latest.csv"


def _write_clay_table_snapshot(root: Path, table_id: str, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    snapshot_rows = [
        {_normalize_key(key): value for key, value in row.items() if value}
        for row in rows
    ]
    _write_csv_rows(_clay_table_snapshot_path(root, table_id), snapshot_rows)


def _save_imported_row_ids(path: Path, imported_row_ids: set[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "imported_clay_row_ids": sorted(imported_row_ids),
    }
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _scalar_payload_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Mapping):
        for key in ("value", "text", "name", "email", "phone", "url", "domain"):
            if key in value:
                scalar = _scalar_payload_value(value[key])
                if scalar:
                    return scalar
        return json.dumps(dict(value), sort_keys=True)
    if isinstance(value, list):
        return ", ".join(part for item in value if (part := _scalar_payload_value(item)))
    return _clean(value)


def _stringify_payload_row(row: Mapping[str, object]) -> dict[str, str]:
    cells: dict[str, str] = {}
    for key, value in row.items():
        scalar = _scalar_payload_value(value)
        if scalar:
            cells[_normalize_header(key)] = scalar
    return cells


def _extract_clay_webhook_rows(payload: object) -> list[dict[str, str]]:
    if isinstance(payload, list):
        rows: list[dict[str, str]] = []
        for item in payload:
            rows.extend(_extract_clay_webhook_rows(item))
        return rows

    if not isinstance(payload, Mapping):
        raise ValueError("Clay webhook payload must be a JSON object or array.")

    row_source: Mapping[str, object] = payload
    for key in ("row", "record", "data", "body", "payload", "values", "columns", "cells"):
        nested = payload.get(key)
        if isinstance(nested, Mapping):
            row_source = nested
            break

    cells = _stringify_payload_row(row_source)
    for source_key in ("id", "row_id", "rowId", "record_id", "recordId"):
        value = payload.get(source_key)
        if value and "clay row id" not in cells:
            cells["clay row id"] = _scalar_payload_value(value)
            break
    return [cells] if any(cells.values()) else []


def _looks_unresolved_clay_token(value: str) -> bool:
    stripped = value.strip()
    return (
        stripped.startswith("/")
        or stripped.startswith("{{")
        or stripped.endswith("}}")
        or bool(re.fullmatch(r"\{[^{}]+\}", stripped))
    )


def _row_hash_identifier(cells: dict[str, str]) -> str:
    material = json.dumps(cells, sort_keys=True).encode("utf-8")
    return hashlib.sha1(material).hexdigest()


def _local_export_row_identifier(row: dict[str, str]) -> str:
    stable_parts = [
        row.get("clay_row_id", ""),
        row.get("source_row_id", ""),
        row.get("source_file", ""),
        row.get("source_row_number", ""),
        row.get("company_name", ""),
        row.get("person_name", ""),
        row.get("email", ""),
    ]
    material = "|".join(part.strip().lower() for part in stable_parts if part)
    if not material:
        material = json.dumps(row, sort_keys=True)
    return hashlib.sha1(material.encode("utf-8")).hexdigest()


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as csv_file:
        return [
            {key: value or "" for key, value in row.items() if key}
            for row in csv.DictReader(csv_file)
        ]


def _write_csv_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    preferred = [
        "local_export_row_id",
        "clay_row_id",
        "source_file",
        "source_row_number",
        "conference",
        "rep_name",
        "rep_email",
        "rep_pipedrive_user_id",
        "company_name",
        "person_name",
        "first_name",
        "last_name",
        "email",
        "phone",
        "linkedin_url",
        "job_title",
        "website_url",
        "domain",
        "sales_leader_name",
        "sales_leader_email",
        "sales_leader_phone",
        "sales_leader_linkedin_url",
        "sales_leader_title",
        "source_notes",
        "last_received_at",
    ]
    all_fields = {key for row in rows for key in row}
    fieldnames = [field for field in preferred if field in all_fields]
    fieldnames.extend(sorted(all_fields - set(fieldnames)))
    with path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_completed_clay_webhook_payload_to_csv(
    payload: object,
    *,
    root: Path | None = None,
    output_path: Path | None = None,
    limit: int | None = None,
) -> dict[str, object]:
    root = setup_ingestion_dirs(root)
    output_path = output_path or _clay_local_export_path(root)
    raw_rows = _extract_clay_webhook_rows(payload)
    if limit is not None:
        raw_rows = raw_rows[:limit]

    existing_rows = _read_csv_rows(output_path)
    rows_by_id = {
        row.get("local_export_row_id") or _local_export_row_identifier(row): row
        for row in existing_rows
    }
    order = [
        row.get("local_export_row_id") or _local_export_row_identifier(row)
        for row in existing_rows
    ]
    exported_rows = 0
    updated_rows = 0
    received_at = datetime.now(timezone.utc).isoformat()

    for cells in raw_rows:
        unresolved = [
            key
            for key, value in cells.items()
            if value and _looks_unresolved_clay_token(value)
        ]
        if unresolved:
            raise ValueError(
                "Clay webhook sent unresolved column placeholders: "
                + ", ".join(sorted(unresolved)[:8])
            )
        row = {_normalize_key(key): value for key, value in cells.items() if value}
        row_id = _local_export_row_identifier(row)
        row["local_export_row_id"] = row_id
        row["last_received_at"] = received_at
        if row_id in rows_by_id:
            rows_by_id[row_id] = {**rows_by_id[row_id], **row}
            updated_rows += 1
        else:
            rows_by_id[row_id] = row
            order.append(row_id)
            exported_rows += 1

    rows = [rows_by_id[row_id] for row_id in order if row_id in rows_by_id]
    _write_csv_rows(output_path, rows)
    return {
        "exported_rows": exported_rows,
        "updated_rows": updated_rows,
        "total_rows": len(rows),
        "csv_path": str(output_path),
    }


def _is_clay_row_ready(
    cells: dict[str, str],
    *,
    ready_column: str,
    ready_value: str,
    ready_any_value: bool,
) -> bool:
    normalized = {_normalize_header(key): value for key, value in cells.items()}
    column_key = _normalize_header(ready_column)
    value = normalized.get(column_key, "").strip()
    if ready_any_value:
        return bool(value)
    return value.lower() == ready_value.strip().lower()


def _fetch_clay_field_name_map(table_id: str) -> dict[str, str]:
    from app.providers import _clay_request

    try:
        _status_code, body = _clay_request("GET", f"/tables/{table_id}", timeout=30.0)
    except httpx.HTTPError:
        return {}

    if not isinstance(body, dict):
        return {}
    table = body.get("table")
    if not isinstance(table, dict):
        return {}
    fields = table.get("fields")
    field_names: dict[str, str] = {}
    if isinstance(fields, list):
        for field in fields:
            if not isinstance(field, dict):
                continue
            field_id = _clean(field.get("id"))
            field_name = _clean(field.get("name"))
            if field_id and field_name:
                field_names[field_id] = field_name
    elif isinstance(fields, dict):
        for field_id, field in fields.items():
            if not isinstance(field, dict):
                continue
            field_name = _clean(field.get("name"))
            if field_id and field_name:
                field_names[str(field_id)] = field_name
    return field_names


def _fetch_clay_table_raw_rows(table_id: str, *, view_id: str = "") -> list[dict[str, str]]:
    from app.providers import _clay_request, _extract_count, _extract_records, _flatten_clay_cells

    try:
        _status_code, count_payload = _clay_request("GET", f"/tables/{table_id}/count", timeout=30.0)
        expected_total = _extract_count(count_payload)
    except httpx.HTTPError:
        expected_total = None

    records: list[dict[str, object]] = []
    page = 1
    records_path = (
        f"/tables/{table_id}/views/{view_id}/records"
        if view_id
        else f"/tables/{table_id}/records"
    )
    while True:
        _status_code, body = _clay_request(
            "GET",
            records_path,
            params={"page": page, "pageSize": CLAY_PULL_PAGE_SIZE},
            timeout=60.0,
        )
        batch = _extract_records(body)
        if not batch:
            break
        records.extend(batch)
        if len(batch) < CLAY_PULL_PAGE_SIZE:
            break
        if expected_total is not None and len(records) >= expected_total:
            break
        page += 1
        if page > 40:
            break

    field_names = _fetch_clay_field_name_map(table_id)
    raw_rows: list[dict[str, str]] = []
    for record in records:
        cells = _flatten_clay_cells(record)
        mapped_cells: dict[str, str] = {}
        for key, value in cells.items():
            mapped_key = field_names.get(key, key)
            mapped_cells[mapped_key] = value
            if mapped_key != key:
                mapped_cells[key] = value
        clay_row_id = str(record.get("id") or record.get("record_id") or record.get("recordId") or "")
        if clay_row_id:
            mapped_cells["clay_row_id"] = clay_row_id
        raw_rows.append({_normalize_header(key): _clean(value) for key, value in mapped_cells.items()})
    return raw_rows


def import_ready_clay_table_rows(
    table_id: str,
    *,
    root: Path | None = None,
    dry_run: bool = True,
    limit: int | None = None,
    ready_column: str | None = None,
    ready_value: str | None = None,
    ready_any_value: bool = False,
    view_id: str | None = None,
    client: PipedriveClient | None = None,
) -> OperationSummary:
    if not table_id:
        return OperationSummary()

    settings = get_settings()
    if not settings.clay_session_cookie:
        return OperationSummary()
    root = setup_ingestion_dirs(root)
    ready_column = ready_column or settings.clay_row_status_column
    ready_value = ready_value or settings.clay_ready_status_value
    view_id = view_id if view_id is not None else settings.trade_show_clay_view_id
    imported_row_ids = _load_imported_clay_rows(root)
    try:
        raw_rows = _fetch_clay_table_raw_rows(table_id, view_id=view_id or "")
    except httpx.HTTPError as exc:
        return OperationSummary(errors=[f"Clay table {table_id}: {exc}"])
    summary = OperationSummary()
    pending_rows: list[dict[str, str]] = []
    pending_row_ids: list[str] = []

    for cells in raw_rows:
        unresolved = [
            key
            for key, value in cells.items()
            if value and _looks_unresolved_clay_token(value)
        ]
        if unresolved:
            raise ValueError(
                "Clay webhook sent unresolved column placeholders: "
                + ", ".join(sorted(unresolved)[:8])
            )
        clay_row_id = _pick(cells, "clay_row_id") or _row_hash_identifier(cells)
        if clay_row_id in imported_row_ids:
            summary.skipped_rows += 1
            continue
        if not _is_clay_row_ready(
            cells,
            ready_column=ready_column,
            ready_value=ready_value,
            ready_any_value=ready_any_value,
        ):
            summary.skipped_rows += 1
            continue
        pending_rows.append(cells)
        pending_row_ids.append(clay_row_id)
        if limit is not None and len(pending_rows) >= limit:
            break

    if not pending_rows:
        return summary

    if not dry_run:
        _write_clay_table_snapshot(root, table_id, pending_rows)

    parsed = normalize_rows(Path(f"clay_table_{table_id}.csv"), pending_rows, limit=limit)
    pipedrive_result = import_rows_to_pipedrive(
        parsed.rows,
        parsed.event,
        dry_run=dry_run,
        client=client,
    )
    summary = merge_summaries(summary, pipedrive_result)
    if not dry_run and not pipedrive_result.errors:
        imported_row_ids.update(pending_row_ids)
        _save_imported_clay_rows(root, imported_row_ids)
    return summary


def import_completed_clay_webhook_payload(
    payload: object,
    *,
    root: Path | None = None,
    dry_run: bool = True,
    limit: int | None = None,
    client: PipedriveClient | None = None,
) -> OperationSummary:
    root = setup_ingestion_dirs(root)
    raw_rows = _extract_clay_webhook_rows(payload)
    if limit is not None:
        raw_rows = raw_rows[:limit]

    imported_row_ids = _load_imported_clay_webhook_rows(root)
    pending_rows: list[dict[str, str]] = []
    pending_row_ids: list[str] = []
    summary = OperationSummary()

    for cells in raw_rows:
        unresolved = [
            key
            for key, value in cells.items()
            if value and _looks_unresolved_clay_token(value)
        ]
        if unresolved:
            raise ValueError(
                "Clay webhook sent unresolved column placeholders: "
                + ", ".join(sorted(unresolved)[:8])
            )
        clay_row_id = _pick(cells, "clay_row_id") or _row_hash_identifier(cells)
        if clay_row_id in imported_row_ids:
            summary.skipped_rows += 1
            continue
        pending_rows.append(cells)
        pending_row_ids.append(clay_row_id)

    if not pending_rows:
        return summary

    parsed = normalize_rows(Path("clay_completed_webhook.csv"), pending_rows, limit=limit)
    pipedrive_result = import_rows_to_pipedrive(
        parsed.rows,
        parsed.event,
        dry_run=dry_run,
        client=client,
        create_notes=False,
    )
    summary = merge_summaries(summary, pipedrive_result)
    if not dry_run and not pipedrive_result.errors:
        imported_row_ids.update(pending_row_ids)
        _save_imported_clay_webhook_rows(root, imported_row_ids)
    return summary


def import_enriched_file(
    path: Path,
    *,
    dry_run: bool = True,
    limit: int | None = None,
    client: PipedriveClient | None = None,
) -> OperationSummary:
    parsed = normalize_sheet(path, limit=limit)
    return import_rows_to_pipedrive(parsed.rows, parsed.event, dry_run=dry_run, client=client)


def summary_to_dict(summary: OperationSummary) -> dict[str, object]:
    return {
        "scanned_files": summary.scanned_files,
        "parsed_rows": summary.parsed_rows,
        "sent_to_clay": summary.sent_to_clay,
        "imported_to_pipedrive": summary.imported_to_pipedrive,
        "updated_people": summary.updated_people,
        "created_people": summary.created_people,
        "updated_orgs": summary.updated_orgs,
        "created_orgs": summary.created_orgs,
        "updated_deals": summary.updated_deals,
        "created_deals": summary.created_deals,
        "created_notes": summary.created_notes,
        "skipped_rows": summary.skipped_rows,
        "errors": summary.errors,
    }


def summary_to_text(summary: OperationSummary) -> str:
    data = summary_to_dict(summary)
    return json.dumps(data, indent=2, sort_keys=True)
