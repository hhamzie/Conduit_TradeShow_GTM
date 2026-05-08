from __future__ import annotations

from collections import Counter
from io import BytesIO
import re

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import Show, ShowGuideRow
from app.show_guides import GUIDE_SHEETS, GuideSheetDefinition, normalize_guide_values, serialize_guide_values
from app.show_intelligence import _load_company_rows


TOKEN_NORMALIZER = re.compile(r"[^a-z0-9]+")


def create_guide_row(
    db: Session,
    *,
    show: Show,
    sheet_key: str,
    payload: dict[str, object],
) -> ShowGuideRow:
    values = normalize_guide_values(sheet_key, payload)
    next_position = db.scalar(
        select(func.coalesce(func.max(ShowGuideRow.position), -1) + 1).where(
            ShowGuideRow.show_id == show.id,
            ShowGuideRow.sheet_key == sheet_key,
        )
    )
    row = ShowGuideRow(
        show=show,
        sheet_key=sheet_key,
        position=int(next_position or 0),
        values_json=serialize_guide_values(values),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def update_guide_row(
    db: Session,
    *,
    row: ShowGuideRow,
    payload: dict[str, object],
) -> None:
    values = normalize_guide_values(row.sheet_key, payload)
    row.values_json = serialize_guide_values(values)
    db.commit()


def delete_guide_row(db: Session, *, row: ShowGuideRow) -> None:
    db.delete(row)
    db.commit()


def _booth_category(booth_number: str) -> str:
    digits = "".join(character for character in booth_number if character.isdigit())
    if len(digits) >= 2:
        return f"{digits[:-2]}00s"
    if digits:
        return f"{digits}00s"
    return "Unassigned"


def _normalize_token(value: object) -> str:
    return TOKEN_NORMALIZER.sub("", str(value or "").strip().lower())


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sheet_key_for_title(title: str) -> str:
    normalized = _normalize_token(title)
    for key, definition in GUIDE_SHEETS.items():
        if normalized in {_normalize_token(key), _normalize_token(definition.label)}:
            return key
    return ""


def _resolve_header_indexes(definition: GuideSheetDefinition, header_row: tuple[object, ...]) -> dict[str, int]:
    normalized_headers = {_normalize_token(value): index for index, value in enumerate(header_row) if value is not None}
    indexes: dict[str, int] = {}
    for field in definition.fields:
        candidates = {_normalize_token(field.key), _normalize_token(field.label)}
        match_index = next((normalized_headers[candidate] for candidate in candidates if candidate in normalized_headers), None)
        if match_index is None:
            raise ValueError(f"{definition.label} is missing the '{field.label}' column.")
        indexes[field.key] = match_index
    return indexes


def import_trade_show_guide_workbook(db: Session, *, show: Show, workbook_bytes: bytes) -> dict[str, int]:
    if not workbook_bytes:
        raise ValueError("Upload an Excel workbook first.")

    try:
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("The uploaded file is not a valid Excel workbook.") from exc

    imported_rows: dict[str, list[dict[str, str]]] = {key: [] for key in GUIDE_SHEETS}
    recognized_sheet_count = 0

    for worksheet in workbook.worksheets:
        sheet_key = _sheet_key_for_title(worksheet.title)
        if not sheet_key:
            continue
        recognized_sheet_count += 1
        definition = GUIDE_SHEETS[sheet_key]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_indexes = _resolve_header_indexes(definition, rows[0])
        for row in rows[1:]:
            payload = {
                field.key: _stringify_cell(row[header_indexes[field.key]]) if header_indexes[field.key] < len(row) else ""
                for field in definition.fields
            }
            if not any(payload.values()):
                continue
            imported_rows[sheet_key].append(normalize_guide_values(sheet_key, payload))

    if recognized_sheet_count == 0:
        raise ValueError("Workbook must include a 'Company Summary' or 'Booth Category Groups' sheet.")

    for row in list(show.guide_rows):
        db.delete(row)
    db.flush()

    for sheet_key, rows in imported_rows.items():
        for position, payload in enumerate(rows):
            db.add(
                ShowGuideRow(
                    show=show,
                    sheet_key=sheet_key,
                    position=position,
                    values_json=serialize_guide_values(payload),
                )
            )
    db.commit()

    return {sheet_key: len(rows) for sheet_key, rows in imported_rows.items()}


def rebuild_trade_show_guides(db: Session, *, show: Show) -> tuple[int, int]:
    company_rows = _load_company_rows(show.latest_export_path)
    if not company_rows:
        raise ValueError("Run the show scrape first so there is export data to build the guide from.")

    for row in list(show.guide_rows):
        db.delete(row)
    db.flush()

    category_counts = Counter(_booth_category(row["booth_number"]) for row in company_rows)
    company_summary_rows: list[ShowGuideRow] = []
    booth_group_rows: list[ShowGuideRow] = []
    for index, company in enumerate(company_rows):
        booth_category = _booth_category(company["booth_number"])
        shared_values = {
            "company_name": company["company_name"],
            "booth_number": company["booth_number"],
            "booth_category": booth_category,
            "sales_team_size": "",
            "customer_service_team_size": "",
            "total_team_size": "",
            "catalog_complexity": "",
            "sales_leader_name": "",
            "sales_leader_role": "",
            "sales_leader_email": "",
            "sales_leader_linkedin": "",
            "source_url": company["website_url"] or show.source_url,
        }
        company_summary_rows.append(
            ShowGuideRow(
                show=show,
                sheet_key="company_summary",
                position=index,
                values_json=serialize_guide_values(shared_values),
            )
        )
        booth_group_rows.append(
            ShowGuideRow(
                show=show,
                sheet_key="booth_category_groups",
                position=index,
                values_json=serialize_guide_values(
                    {
                        **shared_values,
                        "category_total_team_size": str(category_counts[booth_category]),
                    }
                ),
            )
        )

    for row in company_summary_rows + booth_group_rows:
        db.add(row)
    db.commit()
    return len(company_summary_rows), len(booth_group_rows)
